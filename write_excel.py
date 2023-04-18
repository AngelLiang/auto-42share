import openpyxl
from datetime import datetime


def write_to_excel(title, url):
    # 生成文件名
    now = datetime.now()
    filename = now.strftime("%Y-%m-%d") + ".xlsx"

    # 打开Excel文件，如果不存在则创建一个新的
    try:
        wb = openpyxl.load_workbook(filename)
    except:
        wb = openpyxl.Workbook()

    # 选择第一个工作表
    sheet = wb.active

    # 找到第一个空行，即最后一行的下一行
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1).value = title
    sheet.cell(row=row, column=2).value = url

    try:
        # 保存文件
        wb.save(filename)
    except PermissionError:
        print('无法写入到excel文件，请检查文件')
        exit(1)
